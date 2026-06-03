from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework import generics, permissions, serializers
from rest_framework.exceptions import ValidationError
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.auditlog.models import AuditLogEntry
from apps.auditlog.services import log_audit_event
from apps.crm.models import CRMCompany
from apps.crm.serializers import CRMCompanySerializer
from apps.pipelines.access import (
    pipelines_with_company_visibility_queryset,
    pipelines_with_contact_visibility_queryset,
    pipelines_with_deal_visibility_queryset,
    user_can_manage_pipeline_companies,
)
from config.pagination import StandardResultsSetPagination


def tenant_company_ids_for_user(user):
    ids = list(user.companies.values_list("id", flat=True))
    if user.company_id and user.company_id not in ids:
        ids.append(user.company_id)
    return ids


def resolve_default_tenant_company(user):
    if user.company_id:
        return user.company

    first_company = user.companies.order_by("name", "id").first()
    if first_company:
        return first_company

    raise serializers.ValidationError({"detail": "This user is not assigned to a tenant company."})


def crm_companies_queryset_for_user(user):
    queryset = CRMCompany.objects.prefetch_related("contact_links__contact").all()
    if getattr(user, "is_platform_admin", False):
        return queryset.filter(tenant_company_id__in=tenant_company_ids_for_user(user))
    queryset = queryset.filter(tenant_company_id__in=tenant_company_ids_for_user(user))
    if getattr(user, "is_company_admin", False) or user.has_app_permission("crm_companies.view"):
        return queryset
    company_visible_pipelines = pipelines_with_company_visibility_queryset(user)
    contact_visible_pipelines = pipelines_with_contact_visibility_queryset(user)
    deal_visible_pipelines = pipelines_with_deal_visibility_queryset(user)
    return queryset.filter(
        Q(contact_links__pipeline__in=company_visible_pipelines)
        | Q(contacts__pipeline__in=company_visible_pipelines)
        | Q(deals__pipeline__in=company_visible_pipelines)
        | Q(contact_links__pipeline__in=contact_visible_pipelines)
        | Q(contacts__pipeline__in=contact_visible_pipelines)
        | Q(deals__pipeline__in=deal_visible_pipelines)
    ).distinct()


def parse_export_list_param(raw_value):
    return [item.strip() for item in str(raw_value or "").split(",") if item.strip()]


def parse_export_boolean(raw_value):
    return str(raw_value or "").strip().lower() in {"1", "true", "yes", "on"}


def apply_export_date_filter(queryset, *, field_lookup, operator, date_from_value, date_to_value):
    if not operator:
        return queryset

    supported_operators = {"before", "after", "between"}
    if operator not in supported_operators:
        raise ValidationError({"detail": "The selected export date filter is invalid."})

    date_from = parse_date(date_from_value) if date_from_value else None
    date_to = parse_date(date_to_value) if date_to_value else None

    if operator in {"before", "after"} and not date_from:
        raise ValidationError({"detail": "Please choose a comparison date for the export filter."})

    if operator == "between" and (not date_from or not date_to):
        raise ValidationError({"detail": "Please choose both start and end dates for the export filter."})

    if operator == "before":
        return queryset.filter(**{f"{field_lookup}__lt": date_from})

    if operator == "after":
        return queryset.filter(**{f"{field_lookup}__gt": date_from})

    if date_from > date_to:
        raise ValidationError({"detail": "The export start date must be before the end date."})
    return queryset.filter(**{f"{field_lookup}__range": (date_from, date_to)})


def build_excel_response(workbook, *, filename):
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    workbook.save(response)
    return response


def format_export_sheet(sheet):
    header_fill = PatternFill(fill_type="solid", fgColor="C89A2B")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            cell_length = len(str(cell.value or ""))
            if cell_length > max_length:
                max_length = cell_length
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 42)

    sheet.freeze_panes = "A2"


class CRMCompanyListCreateView(generics.ListCreateAPIView):
    serializer_class = CRMCompanySerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = crm_companies_queryset_for_user(self.request.user)
        search = self.request.query_params.get("search", "").strip()
        industry = self.request.query_params.get("industry", "").strip()
        has_contacts = self.request.query_params.get("has_contacts", "").strip().lower()

        if search:
            queryset = queryset.filter(
                Q(name__icontains=search)
                | Q(industry__icontains=search)
                | Q(owner_name__icontains=search)
                | Q(email__icontains=search)
                | Q(website__icontains=search)
                | Q(linkedin_url__icontains=search)
                | Q(address__icontains=search)
                | Q(address_country__icontains=search)
                | Q(address_state__icontains=search)
                | Q(address_line__icontains=search)
                | Q(phone_number__icontains=search)
                | Q(contact_links__contact__full_name__icontains=search)
                | Q(contact_links__contact__email__icontains=search)
                | Q(contact_links__contact__phone__icontains=search)
            ).distinct()

        if industry:
            queryset = queryset.filter(industry=industry)

        if has_contacts == "yes":
            queryset = queryset.filter(contact_links__isnull=False).distinct()
        elif has_contacts == "no":
            queryset = queryset.filter(contact_links__isnull=True)

        return queryset

    def perform_create(self, serializer):
        if not self.request.user.has_app_permission("crm_companies.create"):
            if not pipelines_with_company_visibility_queryset(self.request.user).filter(
                Q(created_by=self.request.user) | Q(memberships__user=self.request.user, memberships__has_full_access=True) | Q(memberships__user=self.request.user, memberships__can_manage_companies=True)
            ).exists():
                raise serializers.ValidationError({"detail": "You do not have permission to create CRM companies."})
        company = serializer.save(tenant_company=resolve_default_tenant_company(self.request.user))
        log_audit_event(
            self.request.user,
            event_type=AuditLogEntry.TYPE_COMPANY,
            action=AuditLogEntry.ACTION_CREATE,
            title="Created CRM company",
            description=company.name,
            target=company,
            company=company.tenant_company,
        )


class CRMCompanyDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = CRMCompanySerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return crm_companies_queryset_for_user(self.request.user)

    def _can_manage_company(self, company):
        if self.request.user.is_platform_admin or self.request.user.is_company_admin or self.request.user.has_app_permission("crm_companies.update"):
            return True
        manageable_pipelines = pipelines_with_company_visibility_queryset(self.request.user)
        for pipeline in manageable_pipelines:
            if user_can_manage_pipeline_companies(self.request.user, pipeline) and (
                company.contact_links.filter(pipeline=pipeline).exists()
                or company.contacts.filter(pipeline=pipeline).exists()
                or company.deals.filter(pipeline=pipeline).exists()
            ):
                return True
        return False

    def perform_update(self, serializer):
        if not self._can_manage_company(self.get_object()):
            raise serializers.ValidationError({"detail": "You do not have permission to update this company."})
        company = serializer.save()
        log_audit_event(
            self.request.user,
            event_type=AuditLogEntry.TYPE_COMPANY,
            action=AuditLogEntry.ACTION_UPDATE,
            title="Updated CRM company",
            description=company.name,
            target=company,
            company=company.tenant_company,
        )

    def perform_destroy(self, instance):
        if not self._can_manage_company(instance):
            raise serializers.ValidationError({"detail": "You do not have permission to delete this company."})
        company_name = instance.name
        tenant_company = instance.tenant_company
        company_ref = instance
        instance.delete()
        log_audit_event(
            self.request.user,
            event_type=AuditLogEntry.TYPE_COMPANY,
            action=AuditLogEntry.ACTION_DELETE,
            title="Deleted CRM company",
            description=company_name,
            target=company_ref,
            company=tenant_company,
        )


class CRMCompanyExportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        queryset = crm_companies_queryset_for_user(request.user).prefetch_related(
            "contact_links__pipeline",
            "contact_links__contact",
            "contacts__pipeline",
            "deals__pipeline",
        )
        export_all = parse_export_boolean(request.query_params.get("export_all", "true"))

        if not export_all:
            pipeline_ids = [int(value) for value in parse_export_list_param(request.query_params.get("pipeline_ids")) if value.isdigit()]
            stage_keys = parse_export_list_param(request.query_params.get("stage_keys"))
            date_field = (request.query_params.get("date_field") or "").strip()
            date_operator = (request.query_params.get("date_operator") or "").strip()
            date_from_value = (request.query_params.get("date_from") or "").strip()
            date_to_value = (request.query_params.get("date_to") or "").strip()

            if pipeline_ids:
                visible_pipeline_ids = set(
                    list(pipelines_with_company_visibility_queryset(request.user).filter(id__in=pipeline_ids).values_list("id", flat=True))
                    + list(pipelines_with_contact_visibility_queryset(request.user).filter(id__in=pipeline_ids).values_list("id", flat=True))
                    + list(pipelines_with_deal_visibility_queryset(request.user).filter(id__in=pipeline_ids).values_list("id", flat=True))
                )
                if visible_pipeline_ids:
                    queryset = queryset.filter(
                        Q(contact_links__pipeline_id__in=visible_pipeline_ids)
                        | Q(contacts__pipeline_id__in=visible_pipeline_ids)
                        | Q(deals__pipeline_id__in=visible_pipeline_ids)
                    ).distinct()
                else:
                    queryset = queryset.none()

            if stage_keys:
                stage_filters = Q()
                for stage_key in stage_keys:
                    pipeline_id, _, status_name = stage_key.partition(":")
                    if pipeline_id.isdigit() and status_name:
                        stage_filters |= (
                            Q(contact_links__pipeline_id=int(pipeline_id), contact_links__status=status_name)
                            | Q(contacts__pipeline_id=int(pipeline_id), contacts__status=status_name)
                            | Q(deals__pipeline_id=int(pipeline_id), deals__stage=status_name)
                        )
                queryset = queryset.filter(stage_filters).distinct() if stage_filters else queryset.none()

            if date_field:
                allowed_date_fields = {
                    "created_at": "created_at__date",
                }
                field_lookup = allowed_date_fields.get(date_field)
                if not field_lookup:
                    raise ValidationError({"detail": "The selected export date field is invalid."})
                queryset = apply_export_date_filter(
                    queryset,
                    field_lookup=field_lookup,
                    operator=date_operator,
                    date_from_value=date_from_value,
                    date_to_value=date_to_value,
                )

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Companies"
        sheet.append(
            [
                "Company",
                "Industry",
                "Owner",
                "Email",
                "Website",
                "Phone numbers",
                "Contact count",
                "Pipelines",
                "Address",
                "Country",
                "State",
                "Employee count",
                "Created date",
            ]
        )

        for company in queryset.order_by("name", "id"):
            pipeline_names = sorted(
                {
                    link.pipeline.name
                    for link in company.contact_links.all()
                    if link.pipeline_id
                }
                | {
                    contact.pipeline.name
                    for contact in company.contacts.all()
                    if contact.pipeline_id
                }
                | {
                    deal.pipeline.name
                    for deal in company.deals.all()
                    if deal.pipeline_id
                }
            )
            sheet.append(
                [
                    company.name,
                    company.industry,
                    company.owner_name,
                    company.email,
                    company.website,
                    " | ".join(company.phone_numbers or ([company.phone_number] if company.phone_number else [])),
                    len(company.contact_links.all()),
                    " | ".join(pipeline_names),
                    company.address,
                    company.address_country,
                    company.address_state,
                    company.employee_count or "",
                    timezone.localtime(company.created_at).date().isoformat() if company.created_at else "",
                ]
            )

        format_export_sheet(sheet)
        return build_excel_response(workbook, filename="companies-export.xlsx")
