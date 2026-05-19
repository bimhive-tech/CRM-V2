import re
from collections import defaultdict

from django.contrib.auth import get_user_model
from django.utils import timezone
from rest_framework.exceptions import ValidationError

from apps.crm.models import CRMCompany, CRMContact, CRMContactCompanyLink


User = get_user_model()

FIELD_ALIASES = {
    "company.name": {"companyname", "company", "organization", "org", "client"},
    "company.website": {"website", "companywebsite", "web", "url", "site"},
    "company.email": {"companyemail", "generalemail", "infoemail"},
    "company.phone": {"companyphone", "tel", "telephone"},
    "company.industry": {"industry", "sector", "type", "category"},
    "contact.name": {"contactname", "contact", "fullname", "name"},
    "contact.first_name": {"firstname", "first", "fname"},
    "contact.last_name": {"lastname", "last", "lname", "surname"},
    "contact.job_title": {"jobtitle", "contactjobtitle", "title", "position", "role"},
    "contact.email": {"contactemail", "contactmail", "personalemail"},
    "contact.phone": {"contactphone", "mobile", "cell", "directline"},
    "status": {"status", "leadstatus", "stage"},
    "sales_person": {"salesperson", "assignedto", "owner", "rep", "agent"},
}

AVAILABLE_IMPORT_FIELDS = [
    {"value": "", "label": "Ignore"},
    {"value": "company.name", "label": "Company name"},
    {"value": "company.website", "label": "Company website"},
    {"value": "company.email", "label": "Company email"},
    {"value": "company.phone", "label": "Company phone"},
    {"value": "company.industry", "label": "Industry"},
    {"value": "contact.name", "label": "Contact name"},
    {"value": "contact.first_name", "label": "Contact first name"},
    {"value": "contact.last_name", "label": "Contact last name"},
    {"value": "contact.job_title", "label": "Contact job title"},
    {"value": "contact.email", "label": "Contact email"},
    {"value": "contact.phone", "label": "Contact phone"},
    {"value": "status", "label": "Status"},
    {"value": "sales_person", "label": "Sales person"},
]

EMAIL_SPLIT_RE = re.compile(r"\s{2,}|\n| / |,")
PHONE_SPLIT_RE = re.compile(r"\s{2,}|\n| / |,| (?<!\d)- | - (?!\d)")


def normalize_header(value):
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower().strip())


def text_value(value):
    if value is None:
        return ""
    return str(value).strip()


def detect_target_field(normalized_header):
    for target, aliases in FIELD_ALIASES.items():
        if normalized_header in aliases:
            return target
    return ""


def header_match_count(row):
    matches = 0
    for value in row:
        if detect_target_field(normalize_header(value)):
            matches += 1
    return matches


def detect_header_row(rows):
    for index, row in enumerate(rows[:5]):
        non_empty = [value for value in row if text_value(value)]
        if len(non_empty) < 3:
            continue
        if header_match_count(row) >= 1:
            return index
    return None


def transpose_rows(rows):
    width = max((len(row) for row in rows), default=0)
    padded = [row + [""] * (width - len(row)) for row in rows]
    return [list(item) for item in zip(*padded)]


def is_transposed(rows):
    if not rows or not rows[0]:
        return False
    first_col = [normalize_header(row[0]) for row in rows[:8] if text_value(row[0])]
    if len(first_col) < 3:
        return False
    recognized = sum(1 for value in first_col if detect_target_field(value))
    return recognized >= 3


def unmerge_sheet(sheet):
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged in merged_ranges:
        top_left = sheet.cell(merged.min_row, merged.min_col).value
        sheet.unmerge_cells(str(merged))
        for row in sheet.iter_rows(merged.min_row, merged.max_row, merged.min_col, merged.max_col):
            for cell in row:
                cell.value = top_left


def forward_fill_company_fields(records):
    last_seen = {}
    keys = [
        "company.name",
        "company.website",
        "company.email",
        "company.phone",
        "company.industry",
    ]
    for record in records:
        if record.get("company.name"):
            last_seen = {key: record.get(key, "") for key in keys}
        else:
            for key, value in last_seen.items():
                if not record.get(key):
                    record[key] = value
    return records


def split_multi_value(value, kind):
    cleaned = text_value(value)
    if not cleaned:
        return []
    parts = (EMAIL_SPLIT_RE if kind == "email" else PHONE_SPLIT_RE).split(cleaned)
    return [part.strip() for part in parts if part and part.strip()]


def extract_primary_email(value):
    for candidate in split_multi_value(value, "email"):
        if "@" in candidate:
            return candidate.strip()
    return ""


def normalize_phone(candidate):
    digits = re.sub(r"\D", "", candidate or "")
    if 7 <= len(digits) <= 15:
        return candidate.strip()
    return ""


def extract_phone_numbers(value):
    seen = []
    for candidate in split_multi_value(value, "phone"):
        phone = normalize_phone(candidate)
        if phone and phone not in seen:
            seen.append(phone)
    return seen


def normalize_status(value):
    raw = text_value(value)
    lowered = raw.lower()
    if any(token in lowered for token in ["sent him mail", "sent her mail", "sent mail", "sent email"]):
        return "email_sent"
    if any(token in lowered for token in ["called him directly", "called directly", "called him", "called her"]):
        return "called"
    if lowered.startswith("meeting") or "meeting" in lowered:
        return "meeting"
    if "in processing" in lowered:
        return "in_progress"
    if any(token in lowered for token in ["waiting to contact", "waiting for appointment", "waiting for quotation"]):
        return "pending"
    return raw


def build_contact_name(record):
    if record.get("contact.name"):
        return record["contact.name"]
    first_name = record.get("contact.first_name", "").strip()
    last_name = record.get("contact.last_name", "").strip()
    return " ".join(part for part in [first_name, last_name] if part).strip()


def parse_sheet(sheet, sheet_index, explicit_mapping=None):
    unmerge_sheet(sheet)
    rows = [[text_value(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    while rows and not any(text_value(value) for value in rows[-1]):
        rows.pop()
    if not rows:
        return None

    if is_transposed(rows):
        rows = transpose_rows(rows)

    header_row_index = detect_header_row(rows)
    if header_row_index is None:
        return {
            "sheet_name": sheet.title,
            "sheet_key": f"sheet_{sheet_index}",
            "requires_manual_mapping": True,
            "columns": [],
            "preview_rows": [],
            "records": [],
        }

    headers = rows[header_row_index]
    normalized_headers = [normalize_header(value) for value in headers]
    columns = []
    for column_index, header in enumerate(headers):
        source_key = f"sheet_{sheet_index}_col_{column_index}"
        suggested_field = explicit_mapping.get(source_key) if explicit_mapping else detect_target_field(normalized_headers[column_index])
        columns.append(
            {
                "source_key": source_key,
                "header": header or f"Column {column_index + 1}",
                "normalized_header": normalized_headers[column_index],
                "suggested_field": suggested_field or "",
            }
        )

    if "contact.name" in [column["suggested_field"] for column in columns]:
        for index, column in enumerate(columns[:-1]):
            next_column = columns[index + 1]
            if column["suggested_field"] == "contact.name" and not next_column["normalized_header"]:
                column["suggested_field"] = "contact.first_name"
                next_column["suggested_field"] = explicit_mapping.get(next_column["source_key"], "contact.last_name") if explicit_mapping else "contact.last_name"

    records = []
    for row in rows[header_row_index + 1 :]:
        if not any(text_value(value) for value in row):
            continue
        mapped = defaultdict(str)
        custom_fields = {}
        for column_index, column in enumerate(columns):
            raw_value = row[column_index] if column_index < len(row) else ""
            target_field = (explicit_mapping or {}).get(column["source_key"], column["suggested_field"])
            if target_field:
                mapped[target_field] = text_value(raw_value)
            elif text_value(raw_value):
                custom_fields[column["header"]] = text_value(raw_value)
        mapped["custom_fields"] = custom_fields
        mapped["source_sheet"] = sheet.title
        records.append(dict(mapped))

    records = forward_fill_company_fields(records)
    normalized_records = []
    for record in records:
        contact_email = extract_primary_email(record.get("contact.email", ""))
        company_email = extract_primary_email(record.get("company.email", ""))
        contact_phone_numbers = extract_phone_numbers(record.get("contact.phone", ""))
        company_phone_numbers = extract_phone_numbers(record.get("company.phone", ""))
        contact_name = build_contact_name(record)

        normalized = {
            "company.name": record.get("company.name", "").strip(),
            "company.website": record.get("company.website", "").strip(),
            "company.email": company_email,
            "company.phone_numbers": company_phone_numbers,
            "company.phone": company_phone_numbers[0] if company_phone_numbers else "",
            "company.industry": record.get("company.industry", "").strip(),
            "contact.name": contact_name,
            "contact.job_title": record.get("contact.job_title", "").strip(),
            "contact.email": contact_email,
            "contact.phone_numbers": contact_phone_numbers,
            "contact.phone": contact_phone_numbers[0] if contact_phone_numbers else "",
            "status": normalize_status(record.get("status", "")),
            "sales_person": record.get("sales_person", "").strip(),
            "custom_fields": record.get("custom_fields", {}),
            "source_sheet": sheet.title,
        }

        if not normalized["company.name"] and not normalized["contact.name"] and not normalized["contact.email"]:
            continue
        if not normalized["company.name"] and not normalized["contact.name"] and not normalized["contact.phone"]:
            continue
        normalized_records.append(normalized)

    preview_rows = [
        {
            "company_name": record["company.name"],
            "contact_name": record["contact.name"],
            "contact_email": record["contact.email"],
            "contact_phone": record["contact.phone"],
            "status": record["status"],
        }
        for record in normalized_records[:5]
    ]

    return {
        "sheet_name": sheet.title,
        "sheet_key": f"sheet_{sheet_index}",
        "requires_manual_mapping": False,
        "columns": columns,
        "preview_rows": preview_rows,
        "records": normalized_records,
    }


def parse_workbook(file_obj, explicit_mapping=None):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as error:
        raise ValidationError(
            {"detail": "Excel import is not available because the server is missing the openpyxl package."}
        ) from error

    workbook = load_workbook(file_obj, data_only=True)
    sheets = []
    all_records = []
    unmatched_columns = []

    for sheet_index, sheet_name in enumerate(workbook.sheetnames):
        result = parse_sheet(workbook[sheet_name], sheet_index, explicit_mapping=explicit_mapping or {})
        if not result:
            continue
        sheets.append(
            {
                "sheet_name": result["sheet_name"],
                "sheet_key": result["sheet_key"],
                "requires_manual_mapping": result["requires_manual_mapping"],
                "columns": result["columns"],
                "preview_rows": result["preview_rows"],
            }
        )
        all_records.extend(result["records"])
        unmatched_columns.extend(
            [
                {"sheet_name": result["sheet_name"], "header": column["header"]}
                for column in result["columns"]
                if not column["suggested_field"]
            ]
        )

    return {
        "sheets": sheets,
        "records": all_records,
        "stats": {
            "sheet_count": len(sheets),
            "row_count": len(all_records),
            "unmatched_column_count": len(unmatched_columns),
        },
        "unmatched_columns": unmatched_columns,
        "available_fields": AVAILABLE_IMPORT_FIELDS,
    }


def resolve_sales_owner(tenant_company, raw_owner_name):
    cleaned = text_value(raw_owner_name)
    if not cleaned:
        return None
    return (
        User.objects.filter(companies=tenant_company, full_name__iexact=cleaned).distinct().first()
        or User.objects.filter(companies=tenant_company, email__iexact=cleaned).distinct().first()
    )


def import_contact_records(records, tenant_company, pipeline=None):
    created_contacts = 0
    created_companies = 0
    created_links = 0
    updated_links = 0
    unresolved_sales_people = []

    imported_at = timezone.now()

    for record in records:
        company_name = text_value(record.get("company.name"))
        contact_email = text_value(record.get("contact.email")).lower()
        contact_name = text_value(record.get("contact.name"))

        if not company_name or not contact_name:
            continue

        company, company_created = CRMCompany.objects.get_or_create(
            tenant_company=tenant_company,
            name=company_name,
            defaults={
                "website": text_value(record.get("company.website")),
                "email": text_value(record.get("company.email")),
                "phone_numbers": record.get("company.phone_numbers", []),
                "phone_number": text_value(record.get("company.phone")),
                "created_by_import": True,
                "imported_at": imported_at,
            },
        )
        if company_created:
            created_companies += 1
        else:
            dirty = False
            for field, value in {
                "website": text_value(record.get("company.website")),
                "email": text_value(record.get("company.email")),
            }.items():
                if value and not getattr(company, field):
                    setattr(company, field, value)
                    dirty = True
            if record.get("company.phone_numbers") and not company.phone_numbers:
                company.phone_numbers = record.get("company.phone_numbers", [])
                dirty = True
            if dirty:
                company.save()

        if contact_email:
            contact = CRMContact.objects.filter(tenant_company=tenant_company, email__iexact=contact_email).first()
        else:
            contact = CRMContact.objects.filter(tenant_company=tenant_company, full_name__iexact=contact_name, company=company).first()

        if contact is None:
            contact = CRMContact.objects.create(
                tenant_company=tenant_company,
                company=company,
                pipeline=pipeline,
                full_name=contact_name,
                title=text_value(record.get("contact.job_title")),
                email=contact_email or f"missing-email-{company.id}-{company.contacts.count() + 1}@placeholder.local",
                phone=text_value(record.get("contact.phone")),
                phone_numbers=record.get("contact.phone_numbers", []),
                status=text_value(record.get("status")) or "Lead",
                created_by_import=True,
                imported_at=imported_at,
            )
            created_contacts += 1
        else:
            dirty = False
            if record.get("contact.phone_numbers") and not contact.phone_numbers:
                contact.phone_numbers = record.get("contact.phone_numbers", [])
                dirty = True
            if text_value(record.get("contact.job_title")) and not contact.title:
                contact.title = text_value(record.get("contact.job_title"))
                dirty = True
            if text_value(record.get("status")) and (not contact.status or contact.status == "Lead"):
                contact.status = text_value(record.get("status"))
                dirty = True
            if pipeline and not contact.pipeline_id:
                contact.pipeline = pipeline
                dirty = True
            if dirty:
                contact.save()

        owner = resolve_sales_owner(tenant_company, record.get("sales_person"))
        if text_value(record.get("sales_person")) and owner is None:
            unresolved_sales_people.append(record.get("sales_person"))

        link, link_created = CRMContactCompanyLink.objects.update_or_create(
            tenant_company=tenant_company,
            contact=contact,
            company=company,
            defaults={
                "pipeline": pipeline,
                "title": text_value(record.get("contact.job_title")),
                "status": text_value(record.get("status")) or "Lead",
                "owner": owner,
                "created_by_import": True,
                "imported_at": imported_at,
            },
        )
        if link_created:
            created_links += 1
        else:
            updated_links += 1

    return {
        "created_contacts": created_contacts,
        "created_companies": created_companies,
        "created_links": created_links,
        "updated_links": updated_links,
        "unresolved_sales_people": sorted({text_value(value) for value in unresolved_sales_people if text_value(value)}),
    }
